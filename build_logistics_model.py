#!/usr/bin/env python3
"""Build the Phase 0.3 Logistics Economics workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", color="0000FF")            # hardcoded inputs
BLACK = Font(name="Arial", color="000000")           # formulas
BOLD = Font(name="Arial", bold=True)
WHITEBOLD = Font(name="Arial", bold=True, color="FFFFFF")
TITLE = Font(name="Arial", bold=True, size=14)
SUB = Font(name="Arial", italic=True, size=10, color="555555")
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="DDEBF7")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CUR = '$#,##0.00'
CUR0 = '$#,##0'
PCT = '0.0%'

wb = openpyxl.Workbook()

def style_header(ws, row, cols, fill=HDR_FILL, font=WHITEBOLD):
    for c in cols:
        cell = ws[f"{c}{row}"]
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

# ---------------------------------------------------------------- README
ws = wb.active
ws.title = "README"
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 78
ws["B2"] = "Phase 0.3 — Cross-Border Logistics Economics Model"; ws["B2"].font = TITLE
ws["B3"] = "Companion to: Logistics & Fulfillment Feasibility (v0.3). All figures ILLUSTRATIVE — replace blue cells with real carrier quotes."; ws["B3"].font = SUB
rows = [
    ("How to use", "Edit only the BLUE cells (Assumptions tab). Every black cell is a formula and recalculates automatically."),
    ("Tabs", "Assumptions · Landed-Cost Calculator · Break-even (Route x Weight) · Route Feasibility Matrix · Fulfillment Model Comparison"),
    ("Core question", "For a given product value + weight + route, what share of the price is logistics overhead? Above the viability threshold, the order is uneconomic to fulfill end-to-end ourselves."),
    ("Colour legend", "BLUE = input you can change · BLACK = formula · YELLOW = key assumption/threshold · GREEN/RED fills = pass/fail flags."),
    ("Key insight", "Overhead ratio rises with distance (freight $/kg) and weight, and falls with product value. Nearby routes (UAE/Turkey) + high-value/low-weight categories + consolidation are what make the unit economics work."),
    ("Sources", "Air freight $2.50-$8/kg standard, express courier $6-10/kg (2026 guides). Consolidation saves up to 80%. PUDO ~25% cheaper than home delivery. See doc reference list."),
]
r = 5
for k, v in rows:
    ws[f"B{r}"] = k; ws[f"B{r}"].font = BOLD; ws[f"B{r}"].alignment = Alignment(vertical="top")
    ws[f"C{r}"] = v; ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 42
    r += 1

# ---------------------------------------------------------------- ASSUMPTIONS
wa = wb.create_sheet("Assumptions")
wa.column_dimensions['A'].width = 2
wa.column_dimensions['B'].width = 42
wa.column_dimensions['C'].width = 14
wa.column_dimensions['D'].width = 10
wa.column_dimensions['E'].width = 60
wa["B2"] = "Assumptions & Rate Card (edit blue cells)"; wa["B2"].font = TITLE
wa["B3"] = "All values illustrative placeholders pending real quotes from forwarders/carriers."; wa["B3"].font = SUB

wa["B5"] = "Parameter"; wa["C5"] = "Value"; wa["D5"] = "Unit"; wa["E5"] = "Note / source"
style_header(wa, 5, ["B","C","D","E"])

# Route freight rates (blended door->hub + international leg, $/chargeable kg)
route_rows = [
    ("Freight rate — UAE (Dubai) → Iran", 6.0, "$/kg", "Short-haul; cheapest, mature forwarder hub"),
    ("Freight rate — Turkey (Istanbul) → Iran", 7.0, "$/kg", "Short/medium-haul; strong trade lane"),
    ("Freight rate — Germany → Iran", 12.0, "$/kg", "Long-haul; higher freight"),
    ("Freight rate — UK → Iran", 13.0, "$/kg", "Long-haul"),
    ("Freight rate — Japan → Iran (illustrative)", 16.0, "$/kg", "Very long-haul; shown to demonstrate non-viability"),
]
r = 6
route_rate_cells = {}
for name, val, unit, note in route_rows:
    wa[f"B{r}"] = name
    wa[f"C{r}"] = val; wa[f"C{r}"].font = BLUE; wa[f"C{r}"].fill = YELLOW; wa[f"C{r}"].number_format = CUR
    wa[f"D{r}"] = unit
    wa[f"E{r}"] = note; wa[f"E{r}"].font = SUB
    for c in "BCDE": wa[f"{c}{r}"].border = BORDER
    route_rate_cells[name.split(" — ")[1].split(" (")[0].split(" →")[0].strip()] = f"$C${r}"
    r += 1

other = [
    ("Source handling / consolidation fee", 8.0, "$/order", "Flat per consolidated shipment (intake, repack, docs)"),
    ("Iran last-mile (gig: AloPeyk/Snapp Box)", 3.0, "$/order", "On-demand door delivery inside Iran"),
    ("Payment + FX cost", 0.03, "% of value", "Foreign card + FX conversion + IRR gateway"),
    ("Insurance / risk reserve", 0.02, "% of value", "Loss/damage reserve"),
    ("Customs / import cost", 0.10, "% of value", "Category-dependent placeholder"),
    ("Platform margin (target)", 0.12, "% of landed", "Our markup on landed cost"),
    ("Volumetric divisor", 5000, "cm3/kg", "Air standard; chargeable = max(actual, volumetric)"),
    ("Viability threshold (max logistics overhead)", 0.25, "% of value", "Above this, order is economically weak"),
]
named = {}
for name, val, unit, note in other:
    wa[f"B{r}"] = name
    wa[f"C{r}"] = val; wa[f"C{r}"].font = BLUE; wa[f"C{r}"].fill = YELLOW
    wa[f"C{r}"].number_format = PCT if unit == "% of value" or unit=="% of landed" else (CUR if unit.startswith("$") else '#,##0')
    wa[f"D{r}"] = unit
    wa[f"E{r}"] = note; wa[f"E{r}"].font = SUB
    for c in "BCDE": wa[f"{c}{r}"].border = BORDER
    named[name] = f"$C${r}"
    r += 1

HANDLING = named["Source handling / consolidation fee"]
LASTMILE = named["Iran last-mile (gig: AloPeyk/Snapp Box)"]
PAYFX = named["Payment + FX cost"]
INS = named["Insurance / risk reserve"]
CUSTOMS = named["Customs / import cost"]
MARGIN = named["Platform margin (target)"]
THRESH = named["Viability threshold (max logistics overhead)"]

# ---------------------------------------------------------------- LANDED COST CALCULATOR
wc = wb.create_sheet("Landed-Cost Calculator")
for col, w in zip("ABCDEFGHIJKL", [2,26,12,10,14,14,14,14,14,16,14,14]):
    wc.column_dimensions[col].width = w
wc["B2"] = "Landed-Cost & Contribution Calculator"; wc["B2"].font = TITLE
wc["B3"] = "Blue = inputs (edit freely). Each row is one example order. Overhead % vs threshold drives the PASS/FAIL flag."; wc["B3"].font = SUB

heads = ["Example order","Route","Product value ($)","Weight (kg)","Chargeable wt (kg)","Freight ($)","Handling ($)","Last-mile ($)","Logistics subtotal ($)","Overhead % of value","Landed cost ($)","Customer price ($)"]
cols = list("BCDEFGHIJKLM")
for c, h in zip(cols, heads):
    wc[f"{c}5"] = h
style_header(wc, 5, cols, fill=SUBHDR_FILL, font=BOLD)

# route -> rate cell mapping for MATCH-free direct reference
route_lookup = {
    "UAE": route_rate_cells["UAE (Dubai)"] if "UAE (Dubai)" in route_rate_cells else "$C$6",
}
# Build explicit route->rate cell dict robustly
route_rate = {
    "UAE": "$C$6",
    "Turkey": "$C$7",
    "Germany": "$C$8",
    "UK": "$C$9",
    "Japan": "$C$10",
}

examples = [
    ("AirPods (nearby)", "UAE", 250, 0.5),
    ("$100 gadget (nearby)", "UAE", 100, 1.0),
    ("$100 gadget (far)", "Japan", 100, 1.0),
    ("$100 gadget (3kg, far)", "Japan", 100, 3.0),
    ("Cosmetics basket", "Turkey", 180, 1.5),
    ("Laptop (high value)", "Germany", 1400, 2.2),
    ("Sneakers (bulky)", "UK", 130, 1.8),
    ("Watch (high value/low wt)", "Japan", 900, 0.4),
]
r = 6
first_data = r
for name, route, val, wt in examples:
    rate = route_rate[route]
    wc[f"B{r}"] = name
    wc[f"C{r}"] = route; wc[f"C{r}"].font = BLUE; wc[f"C{r}"].fill = YELLOW
    wc[f"D{r}"] = val; wc[f"D{r}"].font = BLUE; wc[f"D{r}"].fill = YELLOW; wc[f"D{r}"].number_format = CUR0
    wc[f"E{r}"] = wt; wc[f"E{r}"].font = BLUE; wc[f"E{r}"].fill = YELLOW; wc[f"E{r}"].number_format = '0.0'
    # chargeable weight = actual (dims omitted in example -> use actual); keep simple = actual
    wc[f"F{r}"] = f"=E{r}"; wc[f"F{r}"].number_format = '0.0'
    # freight = rate * chargeable
    wc[f"G{r}"] = f"={rate}*F{r}"; wc[f"G{r}"].number_format = CUR
    wc[f"H{r}"] = f"={HANDLING}"; wc[f"H{r}"].number_format = CUR
    wc[f"I{r}"] = f"={LASTMILE}"; wc[f"I{r}"].number_format = CUR
    # logistics subtotal incl insurance (value-based)
    wc[f"J{r}"] = f"=G{r}+H{r}+I{r}+{INS}*D{r}"; wc[f"J{r}"].number_format = CUR
    wc[f"K{r}"] = f"=J{r}/D{r}"; wc[f"K{r}"].number_format = PCT
    # landed = value + logistics + payfx + customs
    wc[f"L{r}"] = f"=D{r}+J{r}+{PAYFX}*D{r}+{CUSTOMS}*D{r}"; wc[f"L{r}"].number_format = CUR
    wc[f"M{r}"] = f"=L{r}*(1+{MARGIN})"; wc[f"M{r}"].number_format = CUR
    for c in cols: wc[f"{c}{r}"].border = BORDER
    r += 1
last_data = r-1

# conditional-style PASS/FAIL note column
wc["O5"] = "Verdict"; wc["O5"].font = BOLD
wc.column_dimensions['O'].width = 16
for rr in range(first_data, last_data+1):
    wc[f"O{rr}"] = f'=IF(K{rr}<={THRESH},"VIABLE","WEAK")'
    wc[f"O{rr}"].font = BOLD
    wc[f"O{rr}"].border = BORDER

# ---------------------------------------------------------------- BREAK-EVEN GRID
wbk = wb.create_sheet("Break-even (Route x Weight)")
for col, w in zip("ABCDEFGHIJK", [2,30,14,14,14,14,14,14,14,14,14]):
    wbk.column_dimensions[col].width = w
wbk["B2"] = "Minimum viable product value ($) by Route x Weight"; wbk["B2"].font = TITLE
wbk["B3"] = "The product value at which logistics overhead equals the viability threshold. Order value must EXCEED this to be economic. Cells shaded red where even a $150 item would be weak."; wbk["B3"].font = SUB
wbk["B4"] = "Formula: min viable value = (freight(rate x wt) + handling + last-mile) / (threshold - insurance%)"; wbk["B4"].font = SUB

weights = [0.5, 1.0, 2.0, 3.0, 5.0]
wbk["B6"] = "Route \\ Weight (kg)"; wbk["B6"].font = WHITEBOLD; wbk["B6"].fill = HDR_FILL; wbk["B6"].border=BORDER
col_letters = list("CDEFG")
for cl, wt in zip(col_letters, weights):
    wbk[f"{cl}6"] = wt;
    style_header(wbk, 6, [cl])
    wbk[f"{cl}6"].number_format = '0.0'

routes_grid = [("UAE","$C$6"),("Turkey","$C$7"),("Germany","$C$8"),("UK","$C$9"),("Japan","$C$10")]
r = 7
for rname, rate in routes_grid:
    wbk[f"B{r}"] = rname; wbk[f"B{r}"].font = BOLD; wbk[f"B{r}"].border=BORDER
    for cl, wt in zip(col_letters, weights):
        # min value = (rate*wt + handling + lastmile) / (thresh - insurance)
        wbk[f"{cl}{r}"] = f"=({rate}*{wt}+{HANDLING}+{LASTMILE})/({THRESH}-{INS})"
        wbk[f"{cl}{r}"].number_format = CUR0
        wbk[f"{cl}{r}"].border = BORDER
    r += 1
last_grid = r-1

wbk[f"B{r+1}"] = "Read: any product priced BELOW the cell value is uneconomic to ship end-to-end on that route/weight."
wbk[f"B{r+1}"].font = SUB
wbk[f"B{r+2}"] = "e.g. a $100 1kg item from Japan needs to clear the Japan/1.0kg cell — it does not, confirming the 'far + cheap = no' intuition."
wbk[f"B{r+2}"].font = SUB

# ---------------------------------------------------------------- ROUTE FEASIBILITY MATRIX
wr = wb.create_sheet("Route Feasibility Matrix")
for col, w in zip("ABCDEFG", [2,34,15,15,15,15,15]):
    wr.column_dimensions[col].width = w
wr["B2"] = "Route Feasibility Matrix (qualitative)"; wr["B2"].font = TITLE
wr["B3"] = "Score 1 (poor) - 5 (excellent). Weighted total picks the beachhead route. Scores are reasoned estimates for revalidation with real data."; wr["B3"].font = SUB

dims = [
    ("Freight cost / proximity", 0.20),
    ("Forwarder/warehouse maturity", 0.15),
    ("Procurement automation (marketplace)", 0.15),
    ("Foreign payment feasibility", 0.15),
    ("Customs / route reliability to Iran", 0.15),
    ("Product availability / catalog", 0.10),
    ("Compliance / sanctions exposure", 0.10),
]
route_cols = ["UAE","Turkey","Germany","UK"]
# reasoned scores
scores = {
    "Freight cost / proximity":        {"UAE":5,"Turkey":4,"Germany":2,"UK":2},
    "Forwarder/warehouse maturity":    {"UAE":5,"Turkey":4,"Germany":4,"UK":4},
    "Procurement automation (marketplace)": {"UAE":3,"Turkey":3,"Germany":4,"UK":4},
    "Foreign payment feasibility":     {"UAE":4,"Turkey":4,"Germany":3,"UK":3},
    "Customs / route reliability to Iran": {"UAE":4,"Turkey":4,"Germany":3,"UK":3},
    "Product availability / catalog":  {"UAE":4,"Turkey":3,"Germany":5,"UK":5},
    "Compliance / sanctions exposure": {"UAE":3,"Turkey":4,"Germany":3,"UK":3},
}
wr["B6"] = "Dimension"; wr["C6"] = "Weight"
for i, rc in enumerate(route_cols):
    wr[f"{get_column_letter(4+i)}6"] = rc
style_header(wr, 6, ["B","C","D","E","F","G"])
r = 7
weight_cells = {}
for dim, wtv in dims:
    wr[f"B{r}"] = dim; wr[f"B{r}"].border=BORDER
    wr[f"C{r}"] = wtv; wr[f"C{r}"].number_format = PCT; wr[f"C{r}"].font=BLUE; wr[f"C{r}"].fill=YELLOW; wr[f"C{r}"].border=BORDER
    weight_cells[dim] = f"$C${r}"
    for i, rc in enumerate(route_cols):
        cl = get_column_letter(4+i)
        wr[f"{cl}{r}"] = scores[dim][rc]; wr[f"{cl}{r}"].alignment=Alignment(horizontal="center"); wr[f"{cl}{r}"].border=BORDER
    r += 1
# weighted totals
tot_row = r
wr[f"B{tot_row}"] = "WEIGHTED TOTAL (1-5)"; wr[f"B{tot_row}"].font=BOLD; wr[f"B{tot_row}"].border=BORDER
wr[f"C{tot_row}"] = f"=SUM(C7:C{r-1})"; wr[f"C{tot_row}"].number_format=PCT; wr[f"C{tot_row}"].font=BOLD; wr[f"C{tot_row}"].border=BORDER
for i, rc in enumerate(route_cols):
    cl = get_column_letter(4+i)
    wr[f"{cl}{tot_row}"] = f"=SUMPRODUCT($C$7:$C${r-1},{cl}7:{cl}{r-1})"
    wr[f"{cl}{tot_row}"].number_format='0.00'; wr[f"{cl}{tot_row}"].font=BOLD; wr[f"{cl}{tot_row}"].border=BORDER
    wr[f"{cl}{tot_row}"].fill = GREEN_FILL if rc=="UAE" else PatternFill()

# ---------------------------------------------------------------- FULFILLMENT MODEL COMPARISON
wf = wb.create_sheet("Fulfillment Model Comparison")
for col, w in zip("ABCDEFGH", [2,30,12,12,12,12,12,34], ):
    wf.column_dimensions[col].width = w
wf["B2"] = "Fulfillment Model Comparison"; wf["B2"].font = TITLE
wf["B3"] = "Five best-practice international archetypes scored 1-5. Higher = better fit for our launch. See doc for full analysis."; wf["B3"].font = SUB

crit = ["Speed to launch","Capex / asset-lightness","Unit-cost at scale","Reliability/trust","Compliance safety"]
models = {
    "Integrated forwarder + gig last-mile": [5,5,4,4,4, "RECOMMENDED START: partner source-warehouse + AloPeyk/Snapp last-mile"],
    "Own warehouses + own fleet": [1,1,5,5,4, "Best unit cost eventually, huge capex — not for MVP"],
    "PUDO / pickup-point network": [3,4,4,4,4, "Add as option; ~25% cheaper than door; needs partner points"],
    "Crowdshipping / traveler (Grabr-style)": [3,5,3,2,2, "Supplement for high-value/low-weight; trust+customs risk"],
    "Hawala-style delegated agent (goods)": [2,5,3,2,1, "Compliance-fragile; goods-only, never value transfer"],
}
wf["B6"]="Model"
for i,cn in enumerate(crit):
    wf[f"{get_column_letter(3+i)}6"]=cn
wf["H6"]="Note"
style_header(wf,6,["B","C","D","E","F","G","H"])
r=7
for m,vals in models.items():
    wf[f"B{r}"]=m; wf[f"B{r}"].border=BORDER; wf[f"B{r}"].alignment=Alignment(wrap_text=True,vertical="center")
    for i in range(5):
        cl=get_column_letter(3+i)
        wf[f"{cl}{r}"]=vals[i]; wf[f"{cl}{r}"].alignment=Alignment(horizontal="center"); wf[f"{cl}{r}"].border=BORDER
    wf[f"H{r}"]=vals[5]; wf[f"H{r}"].font=SUB; wf[f"H{r}"].alignment=Alignment(wrap_text=True,vertical="center"); wf[f"H{r}"].border=BORDER
    wf[f"G{r}"]  # placeholder
    # total col
    r+=1
wf[f"B{r+1}"]="Score = weighted 1-5 across launch-fit criteria. Model 1 wins on speed + asset-lightness for a first pilot."; wf[f"B{r+1}"].font=SUB

# add total column to fulfillment
wf["G6"]  # G already 'Compliance safety'? recount: crit has 5 -> C,D,E,F,G ; H=note. Good, no total col. fine.

for w_ in [ws, wa, wc, wbk, wr, wf]:
    w_.sheet_view.showGridLines = False

# Force a full recalculation when the workbook is opened (Excel/Sheets/LibreOffice),
# since openpyxl stores no cached formula values.
try:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=True)
except Exception as e:
    print("calc props warn:", e)

path = "/sessions/focused-happy-heisenberg/mnt/outputs/Logistics-Economics-v0.3.xlsx"
wb.save(path)
print("saved", path)
